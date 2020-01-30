import time
from selenium import webdriver

# 브라우저를 실행
d = webdriver.Chrome('./chromedriver')
data = {}
try:
    # 네이버 뉴스 페이지로
    d.get('https://news.naver.com/')

    keys = [
        'right.ranking_tab_100',
        'right.ranking_tab_101',
        'right.ranking_tab_102',
        'right.ranking_tab_103',
        'right.ranking_tab_104',
        'right.ranking_tab_105',
    ]

    for k in keys:
        header = d.find_element_by_id(k)
        header.click()
        # tap 별로 바뀌는 시간을 고려하여 time을 걸어준다.
        time.sleep(0.5)
        data[header.text] = []

        # id 속성값이 right.ranking_contents인 태그를 찾아줘
        elem = d.find_element_by_id('right.ranking_contents')
        lis = elem.find_elements_by_tag_name('li')

        for li in lis:
            atag = li.find_element_by_tag_name('a')
            # 태그 출력
            print(atag.text)
            data[header.text].append(atag.text)

        # time.sleep(2) - 잘되고 있나 확인용

except Exception as e:
    print(e)

finally:
    d.close()
    d.quit()

from openpyxl import Workbook

# Workbook() - 빈 엑셀을 연다
wb = Workbook()

for key, rows in data.items():
    key = key.replace('/', '-')
    ws = wb.create_sheet(key)
    for row in rows:
        ws.append((row,))

wb.save('results.xlsx')